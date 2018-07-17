# xlsx    xlsx

#openpyxl

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    file = load_workbook(filename=path)#所有表格的名称
    print(file.get_sheet_names())
    sheets = file.get_sheet_names()

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)
        # 再将一张表的数据存到字典中
        dic[sheetName] = sheetInfo
    return dic


path = r"E:\1编程\千锋python\Code\46、excel自动化办公\LaGouDataPython.xlsx"
dic = readXlsxFile(path)
